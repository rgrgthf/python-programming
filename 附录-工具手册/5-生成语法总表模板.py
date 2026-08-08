# -*- coding: utf-8 -*-
# ============================================================
# 生成《Python 语法与函数总表》xlsx 模板
# ============================================================
# 用法：python 5-生成语法总表模板.py
# 作用：生成一个带格式的 Excel 模板，分章节分组，供你填写
#       （这个脚本用到的 openpyxl 就是第23章要学的库！）
# 说明：内容由你自己填，表格骨架老师帮你搭好。
# ============================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# ---- 全部章节的语法总表数据（预填完整）----
# 每章：[(语法, 作用, 示例), ...]
DATA = {
"第1章 Git 与 GitHub": [
    ("git init", "初始化仓库", "git init"),
    ("git add 文件", "加入暂存区", "git add 笔记.py"),
    ("git commit -m \"\"", "提交保存", "git commit -m \"第1章\""),
    ("git status", "查看状态", "git status"),
    ("git log", "查看历史", "git log --oneline"),
    ("git branch", "查看分支", "git branch"),
    ("git checkout -b 名", "新建并切换分支", "git checkout -b dev"),
    ("git merge 分支", "合并分支", "git merge dev"),
    ("git push / pull", "推送/拉取", "git push origin main"),
    (".gitignore", "忽略文件", "config.json"),
],
"第2章 输入输出与基础语法": [
    ("print()", "输出到屏幕", "print(\"你好\")"),
    ("print(..., end=\"\")", "不换行输出", "print(\"a\", end=\"\")"),
    ("print(..., sep=\"\")", "指定分隔符", "print(1,2,sep=\"-\")"),
    ("input(\"提示\")", "获取输入(字符串)", "name = input(\"姓名:\")"),
    ("f\"{变量}\"", "格式化字符串", "print(f\"年龄{age}\")"),
    ("int()/float()", "类型转换", "age = int(input(\"年龄:\"))"),
    ("# 注释", "单行注释", "# 说明"),
    ("r\"...\"", "原始字符串", "path = r\"C:\\a\\b\""),
],
"第3章 数据类型与运算符": [
    ("int / float / bool", "数字与布尔", "x = 3; y = 3.14; b = True"),
    ("str", "字符串", "s = \"hello\""),
    ("type()", "查看类型", "type(3)"),
    ("+ - * / // % **", "算术运算", "7 // 2 == 3, 7 % 2 == 1"),
    ("== != > <", "比较运算", "3 > 2"),
    ("and / or / not", "逻辑运算", "a and b"),
    ("& | ^ << >>", "位运算", "1 << 3 == 8"),
    ("= += -= *=", "赋值运算", "x += 1"),
    ("int(\"123\")", "字符串转数字", "int(\"42\")"),
    ("len()", "长度", "len(\"abc\")"),
],
"第4章 流程控制": [
    ("if 条件:", "单分支", "if x > 0: ..."),
    ("if/else", "双分支", "if a: ... else: ..."),
    ("if/elif/else", "多分支", "if a: ... elif b: ..."),
    ("for x in 可迭代:", "遍历循环", "for i in range(5): ..."),
    ("range(起,止,步)", "生成序列", "range(1,10,2)"),
    ("while 条件:", "条件循环", "while x < 10: ..."),
    ("break", "跳出循环", "break"),
    ("continue", "跳过本次", "continue"),
    ("pass", "占位", "pass"),
    ("循环 else", "正常跑完执行", "for...else: ..."),
],
"第5章 数据结构": [
    ("列表 []", "可变有序", "lst = [1,2,3]"),
    ("append/extend", "追加元素", "lst.append(4)"),
    ("insert/remove/pop", "增删元素", "lst.pop()"),
    ("切片 [a:b]", "取子序列", "lst[1:3]"),
    ("元组 ()", "不可变序列", "t = (1,2)"),
    ("字典 {} 键:值", "键值对", "d = {\"a\":1}"),
    ("d.get(k,默认)", "安全取值", "d.get(\"a\", 0)"),
    ("集合 set()", "去重/集合运算", "s = {1,2,3}"),
    ("列表推导式", "快速生成", "[x*x for x in range(5)]"),
    ("in 判断", "成员判断", "\"a\" in lst"),
],
"第6章 函数": [
    ("def 函数名():", "定义函数", "def add(a,b): return a+b"),
    ("参数", "传参", "def f(a, b=1, *args, **kw)"),
    ("return", "返回值", "return 结果"),
    ("作用域", "局部/全局", "global x"),
    ("递归", "函数调自己", "def f(n): return n*f(n-1)"),
    ("lambda", "匿名函数", "f = lambda x: x*2"),
    ("内置函数", "常用工具", "len/sum/max/min/round"),
    ("enumerate()", "带索引遍历", "for i,x in enumerate(lst)"),
    ("zip()", "并行遍历", "for a,b in zip(x,y)"),
],
"第7章 文件与异常处理": [
    ("open(路径,模式)", "打开文件", "f = open(\"a.txt\",\"w\")"),
    ("模式 r/w/a", "读写追加", "open(\"a\",\"r\")"),
    ("read/readline/readlines", "读取", "content = f.read()"),
    ("write()", "写入", "f.write(\"内容\")"),
    ("with open() as f", "自动关闭", "with open(\"a\") as f: ..."),
    ("encoding=\"utf-8\"", "指定编码", "open(\"a\",encoding=\"utf-8\")"),
    ("try/except", "捕获异常", "try: ... except ValueError: ..."),
    ("else / finally", "异常补充", "try...else...finally"),
    ("raise", "主动抛异常", "raise ValueError(\"x\")"),
],
"第8章 面向对象与模块": [
    ("class 类名:", "定义类", "class Dog:"),
    ("def __init__", "构造/初始化", "def __init__(self,name)"),
    ("self", "对象自身", "self.name = name"),
    ("实例属性/方法", "对象的数据与行为", "d.bark()"),
    ("类属性/@classmethod", "类级别", "Tool.count"),
    ("@staticmethod", "静态方法", "@staticmethod def help():"),
    ("继承 class A(B)", "子类继承父类", "class Dog(Animal)"),
    ("super()", "调用父类方法", "super().__init__(name)"),
    ("import 模块", "导入模块", "import math"),
    ("if __name__==\"__main__\"", "主入口", "if __name__ == \"__main__\":"),
],
"第9章 正则表达式": [
    ("re.search(模式,文本)", "找第一个", "re.search(r\"\\d+\", s)"),
    ("re.findall", "找全部", "re.findall(r\"\\d+\", s)"),
    ("re.sub", "替换", "re.sub(r\"\\d\",\"X\",s)"),
    ("re.match/fullmatch", "开头/全匹配", "re.fullmatch(模式,s)"),
    (r"\\d \\w \\s", "数字/单词/空白", "r\"\\d+\""),
    ("量词 * + ? {n}", "重复次数", "r\"a{2,3}\""),
    ("分组 ( )", "提取部分", "re.search(r\"(\\w+)-(\\d+)\",s)"),
    ("命名分组 (?P<n>)", "起名分组", "(?P<name>\\w+)"),
],
"第10章 函数式与推导式": [
    ("lambda", "匿名函数", "lambda x: x*2"),
    ("sorted(key=)", "按规则排序", "sorted(d, key=lambda x:x[1])"),
    ("map(func, iter)", "映射", "list(map(f, lst))"),
    ("filter(func, iter)", "过滤", "list(filter(f, lst))"),
    ("reduce(func, iter)", "累积", "reduce(lambda a,b:a+b, lst)"),
    ("推导式 [x for x in]", "列表推导", "[x for x in r if x>0]"),
    ("生成器表达式 (x for)", "惰性推导", "(x for x in r)"),
    ("闭包", "函数记住外部变量", "def outer(): def inner(): ..."),
],
"第11章 装饰器": [
    ("@decorator", "给函数加功能", "@timer\ndef f(): ..."),
    ("def wrapper(*args)", "包装函数", "def wrapper(*a, **k): ..."),
    ("functools.wraps", "保留元信息", "@wraps(func)"),
    ("带参装饰器", "三层嵌套", "@repeat(3)"),
    ("lru_cache", "缓存", "@lru_cache def f(): ..."),
],
"第12章 生成器与迭代器": [
    ("iter()/next()", "迭代器操作", "next(it)"),
    ("yield", "生成器函数", "def g(): yield 1"),
    ("生成器表达式", "省内存", "(x for x in range(1e6))"),
    ("for line in 文件", "逐行读大文件", "for line in f: ..."),
    ("itertools", "迭代工具", "combinations/permutations"),
],
"第13章 上下文管理器": [
    ("with", "自动管理资源", "with open(\"a\") as f:"),
    ("__enter__/__exit__", "进入/退出", "class T: def __enter__: ..."),
    ("@contextmanager", "简写上下文", "@contextmanager\ndef t(): yield"),
    ("try/finally", "保证清理", "try: ... finally: ..."),
],
"第14章 面向对象进阶": [
    ("@property", "方法变属性", "@property def price(self): ..."),
    ("@xxx.setter", "属性校验", "@price.setter def price(self,v): ..."),
    ("魔法方法 __xx__", "内置操作", "__str__/__eq__/__lt__/__add__"),
    ("ABC/@abstractmethod", "抽象基类", "class A(ABC): @abstractmethod"),
    ("__slots__", "省内存", "__slots__ = [\"x\"]"),
],
"第15章 类型注解与规范": [
    ("参数: 类型", "类型注解", "def f(a: int) -> int:"),
    ("Optional/Union", "可选/联合类型", "def f() -> Optional[int]"),
    ("list[int]", "容器注解", "def f(x: list[float])"),
    ("snake_case/PascalCase", "命名规范", "my_var / MyClass"),
    ("docstring", "文档字符串", "\"\"\"说明\"\"\""),
],
"第16章 标准库宝典": [
    ("Path", "路径操作", "Path(\"a\") / \"b.py\""),
    ("glob()", "通配匹配", "Path(\".\").glob(\"*.py\")"),
    ("os.path", "经典路径", "os.path.join(a,b)"),
    ("datetime", "日期时间", "datetime.now()"),
    ("strftime/strptime", "格式/解析", "now.strftime(\"%Y-%m-%d\")"),
    ("Counter", "计数器", "Counter(列表).most_common()"),
    ("defaultdict", "默认字典", "defaultdict(list)"),
    ("itertools/functools", "工具库", "combinations / partial"),
],
"第17章 调试与日志": [
    ("pdb", "命令行调试", "import pdb; pdb.set_trace()"),
    ("logging", "日志系统", "logging.info(\"x\")"),
    ("日志级别", "分级输出", "DEBUG<INFO<WARNING<ERROR"),
    ("timeit", "测性能", "timeit.timeit(\"代码\")"),
    ("cProfile", "性能剖析", "cProfile.run(\"main()\")"),
],
"第18章 pandas数据分析": [
    ("pd.read_csv()", "读CSV", "df = pd.read_csv(\"a.csv\")"),
    ("DataFrame", "表格数据", "pd.DataFrame(字典)"),
    ("df.head()/shape", "查看数据", "df.head(); df.shape"),
    ("df[\"列\"]", "取列", "df[\"浓度\"]"),
    ("df.loc/iloc", "按标签/位置取", "df.iloc[0,1]"),
    ("df.isna()/fillna()", "处理缺失", "df.fillna(df.mean())"),
    ("df[条件]", "筛选行", "df[df[\"组\"]==\"给药\"]"),
    ("df.groupby().mean()", "分组统计", "df.groupby(\"组\").mean()"),
    ("df.merge/concat", "合并", "pd.concat([df1,df2])"),
    ("df.to_csv()", "导出", "df.to_csv(\"out.csv\", index=False)"),
],
"第19章 numpy科学计算": [
    ("np.array()", "创建数组", "a = np.array([1,2,3])"),
    ("np.arange/linspace", "生成序列", "np.arange(10)"),
    ("shape/dtype/ndim", "数组属性", "a.shape"),
    ("索引/切片", "取元素", "a[1:3], a[行,列]"),
    ("布尔筛选", "条件取数", "a[a > 5]"),
    ("广播", "形状扩展运算", "a + 1"),
    ("np.mean/std/sum", "统计", "a.mean(), a.std()"),
    ("np.dot/@", "矩阵乘法", "A @ B"),
    ("reshape()", "变形", "a.reshape(2,3)"),
],
"第20章 数据可视化": [
    ("plt.plot/scatter", "折线/散点", "plt.plot(x, y)"),
    ("plt.xlabel/ylabel", "坐标轴标签", "plt.xlabel(\"浓度\")"),
    ("plt.title", "标题", "plt.title(\"曲线\")"),
    ("plt.show()/savefig()", "显示/保存", "plt.savefig(\"a.png\",dpi=300)"),
    ("plt.subplots", "多子图", "fig, ax = plt.subplots(1,2)"),
    ("plt.rcParams", "全局设置", "plt.rcParams[\"font.sans-serif\"]"),
    ("sns 统计图", "seaborn绘图", "sns.boxplot(data=df, x=..)"),
],
"第21章 scipy科研统计": [
    ("scipy.stats", "统计模块", "import scipy.stats as st"),
    ("ttest_ind/ttest_rel", "t检验", "st.ttest_ind(a, b)"),
    ("shapiro", "正态性检验", "st.shapiro(x)"),
    ("f_oneway", "单因素ANOVA", "st.f_oneway(a,b,c)"),
    ("chi2_contingency", "卡方检验", "st.chi2_contingency(表)"),
    ("levene", "方差齐性", "st.levene(a, b)"),
],
"第22章 回归与曲线拟合": [
    ("linregress", "线性回归", "st.linregress(x, y)"),
    ("R² (rvalue**2)", "拟合优度", "r2 = result.rvalue**2"),
    ("curve_fit", "非线性拟合", "curve_fit(模型, x, y, p0)"),
    ("标准曲线", "浓度-吸光度", "y = kx + b 反解 x"),
    ("IC50", "半数抑制浓度", "S形拟合取 50%"),
],
"第23章 Excel自动化": [
    ("openpyxl", "读写Excel", "from openpyxl import Workbook"),
    ("load_workbook()", "打开已有", "wb = load_workbook(\"a.xlsx\")"),
    ("wb.active/ws[\"A1\"]", "取表/单元格", "ws[\"A1\"] = 值"),
    ("wb.save()", "保存", "wb.save(\"a.xlsx\")"),
    ("df.to_excel()", "pandas写Excel", "df.to_excel(\"a.xlsx\")"),
    ("ExcelWriter", "多sheet", "with pd.ExcelWriter(\"a\") as w:"),
],
"第24章 测试与代码质量": [
    ("pytest", "测试框架", "def test_xxx():"),
    ("assert", "断言", "assert add(1,2) == 3"),
    ("pytest.approx", "浮点近似", "assert x == approx(0.3)"),
    ("test_ 前缀", "测试命名", "test_add()"),
    ("fixture", "准备数据", "@pytest.fixture"),
],
"第25章 命令行工具与打包": [
    ("argparse", "命令行参数", "parser.add_argument(\"--file\")"),
    ("parse_args()", "解析参数", "args = parser.parse_args()"),
    ("if __name__==\"__main__\"", "入口", "if __name__ == \"__main__\":"),
    ("PyInstaller", "打包exe", "pyinstaller -F 文件.py"),
    ("-F / -w 参数", "单文件/无窗口", "pyinstaller -F -w 脚本.py"),
],
"第26章 网络请求": [
    ("requests.get()", "发起请求", "r = requests.get(url)"),
    ("params=", "传查询参数", "requests.get(url, params={\"q\":1})"),
    ("r.status_code", "状态码", "r.status_code == 200"),
    ("r.json()", "解析JSON", "data = r.json()"),
    ("timeout", "超时", "requests.get(url, timeout=10)"),
    ("headers", "请求头", "headers={\"User-Agent\":\"...\"}"),
],
"第27章 爬虫实战": [
    ("BeautifulSoup", "解析HTML", "soup = BeautifulSoup(html)"),
    ("soup.find/find_all", "找标签", "soup.find_all(\"a\")"),
    ("tag.text / .get()", "取文本/属性", "tag.text, a.get(\"href\")"),
    ("pd.read_html()", "读网页表格", "pd.read_html(html)"),
    ("robots.txt", "爬虫守则", "尊重站点规则"),
],
"第28章 Web开发入门": [
    ("Flask", "Web框架", "from flask import Flask"),
    ("@app.route()", "路由", "@app.route(\"/\")"),
    ("app.run()", "启动服务", "app.run(debug=True)"),
    ("jsonify()", "返回JSON", "return jsonify({\"a\":1})"),
    ("request", "获取请求", "request.args.get(\"q\")"),
],
"第29章 数据库与数据存储": [
    ("sqlite3", "SQLite数据库", "import sqlite3"),
    ("conn.execute()", "执行SQL", "cur.execute(\"SELECT * FROM t\")"),
    ("conn.commit()", "提交事务", "conn.commit()"),
    ("参数化 ?", "防注入", "execute(\"WHERE id=?\", (id,))"),
    ("json.dumps/loads", "序列化", "json.dumps(d)"),
],
"第30章 多线程与多进程": [
    ("threading", "多线程", "import threading"),
    ("ThreadPoolExecutor", "线程池", "with ThreadPoolExecutor() as e:"),
    ("GIL", "解释器锁", "CPU密集用多进程"),
    ("multiprocessing", "多进程", "from multiprocessing import Pool"),
    ("Pool.map()", "并行映射", "pool.map(func, data)"),
],
"第31章 异步编程": [
    ("async def", "协程", "async def main():"),
    ("await", "等待", "await func()"),
    ("asyncio.run()", "启动", "asyncio.run(main())"),
    ("asyncio.sleep", "异步等待", "await asyncio.sleep(1)"),
],
"第32章 RDKit化学信息学": [
    ("Chem.MolFromSmiles", "SMILES→分子", "mol = Chem.MolFromSmiles(s)"),
    ("mol.GetNumAtoms()", "原子数", "mol.GetNumAtoms()"),
    ("Descriptors", "分子描述符", "Descriptors.MolWt(mol)"),
    ("GetMorganFingerprint", "分子指纹", "Morgan 指纹"),
    ("Tanimoto相似度", "相似度比较", "DataStructs.TanimotoSimilarity"),
    ("SDWriter", "写SDF", "SDWriter(\"a.sdf\")"),
],
"第33章 RDKit进阶": [
    ("药效团", "活性基团模式", "SMARTS 搜索"),
    ("3D构象", "AddHs+Embed", "AllChem.EmbedMolecule(mol)"),
    ("分子对接", "Vina", "对接打分"),
    ("SDMolSupplier", "批量读SDF", "SDMolSupplier(\"库.sdf\")"),
],
"第34章 机器学习与QSAR": [
    ("train_test_split", "划分数据", "train_test_split(X, y)"),
    ("模型三步", "建/训/测", "model.fit(X,y); model.predict(X)"),
    ("分类评估", "accuracy/混淆矩阵", "accuracy_score(y, pred)"),
    ("回归评估", "R² / MSE", "r2_score(y, pred)"),
    ("StandardScaler", "标准化", "scaler.fit_transform(X)"),
    ("过拟合", "训练好测试差", "用测试集评估"),
],
"第35章 机器学习深入": [
    ("特征工程", "造有用特征", "标准化/编码/选择"),
    ("cross_val_score", "交叉验证", "cross_val_score(model, X, y)"),
    ("GridSearchCV", "调参", "GridSearchCV(model, 网格)"),
    ("class_weight", "类别平衡", "class_weight=\"balanced\""),
    ("F1/AUC", "不平衡指标", "用 F1/AUC 评估"),
],
"第36章 深度学习入门": [
    ("Tensor", "张量", "torch.tensor([1,2])"),
    ("nn.Module", "定义网络", "class Net(nn.Module)"),
    ("损失/优化器", "训练组件", "nn.MSELoss(); optim.Adam"),
    ("训练循环", "前向+反向", "loss.backward(); opt.step()"),
    ("torch.save", "保存模型", "torch.save(model.state_dict())"),
],
"第37章 药物发现实战": [
    ("虚拟筛选", "库→打分→排序", "粗筛→对接→候选"),
    ("ZINC", "化合物大库", "免费分子库"),
    ("对接打分", "近似结合能", "越负越好"),
    ("湿实验验证", "闭环", "计算结果要验证"),
],
"第38章 综合项目与作品集": [
    ("MVP", "最小可行版本", "先跑通再迭代"),
    ("项目结构", "说明+代码+结果", "README/代码/数据"),
    ("作品集", "能力证据", "2~3个精作品"),
    ("可复现", "记录步骤", "固定seed/参数记录"),
],
"附录 工具手册": [
    ("VS Code", "编辑器", "Ctrl+Shift+P / F5"),
    ("conda", "环境管理", "conda create -n 名"),
    ("Jupyter", "交互笔记", "Shift+Enter 运行"),
    ("files.exclude", "隐藏文件", "视图隐藏≠删除"),
],
}

# ---- 样式定义（VS Code 同款：代码=Consolas，中文=微软雅黑）----
CODE_FONT = "Consolas"              # 英文/代码用（等宽，和 VS Code 代码一样）
CN_FONT = "Microsoft YaHei"         # 中文用（和 VS Code 中文一样，圆润好看）
title_font = Font(bold=True, size=16, name=CN_FONT)
header_font = Font(bold=True, size=12, color="FFFFFF", name=CN_FONT)
header_fill = PatternFill("solid", fgColor="4472C4")     # 表头蓝
chapter_font = Font(bold=True, size=11, color="FFFFFF", name=CN_FONT)
chapter_fill = PatternFill("solid", fgColor="70AD47")    # 章节绿
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ---- 建工作簿 ----
wb = Workbook()
ws = wb.active
ws.title = "语法总表"

# 标题行（合并 A1:C1）
ws.merge_cells("A1:C1")
ws["A1"] = "Python 语法与函数总表（第1~38章 + 附录）"
ws["A1"].font = title_font
ws["A1"].alignment = center
ws.row_dimensions[1].height = 30

# 表头行
headers = ["语法", "作用", "示例"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=2, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border
ws.row_dimensions[2].height = 22

# 逐章：章节分组行 + 预填的语法条目
r = 3
for ch, items in DATA.items():
    # 章节分组行（合并横跨3列）
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(row=r, column=1, value=ch)
    c.font = chapter_font
    c.fill = chapter_fill
    c.alignment = center
    ws.row_dimensions[r].height = 20
    r += 1

    # 该章的语法条目（A语法 / B作用 / C示例）
    for gram, effect, example in items:
        ws.cell(row=r, column=1, value=gram).font = Font(name=CODE_FONT)
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=1).alignment = left
        ws.cell(row=r, column=2, value=effect).font = Font(name=CN_FONT)
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=2).alignment = left
        ws.cell(row=r, column=3, value=example).font = Font(name=CODE_FONT)
        ws.cell(row=r, column=3).border = border
        ws.cell(row=r, column=3).alignment = left
        r += 1

# 列宽（语法 / 作用 / 示例）
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 45
ws.column_dimensions["C"].width = 40

# ---- 保存到本脚本同目录 ----
out = Path(__file__).parent / "Python语法总表.xlsx"
wb.save(out)
print(f"✅ 语法总表已生成（预填完整，共{len(DATA)}个章节）：{out}")
print("接下来：用 Excel/WPS 打开，按章节检索语法，忘了就回来查。")

# ============================================================
# 六、易错点汇总
# ============================================================
# 1. 语法总表是"自查手册"：每章语法要点汇总，
#    忘了回来翻，别背
# 2. 生成的总表用 openpyxl 写入 Excel，方便筛选/搜索
# 3. 总表按章节组织，绿色标题行 = 章节分隔
# 4. 生成的 .xlsx 用 Excel/WPS 打开即可；
#    改表要重跑脚本会覆盖，注意先备份
# 5. 语法总表是"目录"，详细内容在各章笔记里
# 6. 复习时：先翻总表找知识点，再去对应章节细看

# ============================================================
# 七、自测（40%基础 + 40%中等 + 20%挑战）
# ============================================================
# 【基础】
# 1. 语法总表是干什么用的？
# 2. 总表用什么库生成？
# 3. 复习时怎么用总表？
#
# 【中等】
# 4. 运行脚本生成总表并用 Excel 打开。
# 5. 说明重跑脚本覆盖文件的注意事项。
# 6. 给总表加一个新章节标题。
#
# 【挑战】
# 7. 改进脚本：把总表内容和各章笔记关联起来。
# 8. 解释"总表当目录、章节当正文"的学习方法。
