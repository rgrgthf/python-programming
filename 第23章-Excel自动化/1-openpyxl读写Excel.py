# ============================================================
# Excel 自动化 ① — openpyxl 读写 Excel
# ============================================================
# ⚠️ 请在 sci 环境运行（如未装：pip install openpyxl）
# 实验数据大量存在 Excel 里。openpyxl 能读写 .xlsx：
#   不用手动改表格，批量处理、批量填充、自动统计。

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ============================================================
# 一、创建并写入 Excel
# ============================================================
wb = Workbook()              # 新建工作簿
ws = wb.active               # 默认工作表
ws.title = "实验数据"         # 改表名

# 写入单元格
ws["A1"] = "样品编号"
ws["B1"] = "浓度"
ws["C1"] = "吸光度"

# 写入多行数据
data = [
    ["A1", 0.1, 0.048],
    ["A2", 0.25, 0.125],
    ["A3", 0.5, 0.251],
    ["A4", 1.0, 0.51],
]
for row in data:
    ws.append(row)           # 从下一行追加

# 保存
wb.save("实验数据.xlsx")
print("已创建 实验数据.xlsx")

# ============================================================
# 二、读取 Excel
# ============================================================
wb2 = load_workbook("实验数据.xlsx")
ws2 = wb2.active

# 读取单个单元格
print(ws2["A1"].value)       # → 样品编号
print(ws2.cell(row=2, column=1).value)   # → A1

# 读取所有数据
for row in ws2.iter_rows(min_row=2, values_only=True):
    print(row)
# → ('A1', 0.1, 0.048) ...

# 获取行数/列数
print(f"共 {ws2.max_row} 行, {ws2.max_column} 列")

# ============================================================
# 三、工作表管理
# ============================================================
ws3 = wb.create_sheet("第2批数据")      # 新建表
ws3["A1"] = "hello"
print(wb.sheetnames)          # → ['实验数据', '第2批数据']

# 切换/删除
# del wb["第2批数据"]         # 删除表
# ws = wb["实验数据"]         # 按名字取表

# ============================================================
# 四、样式：加粗标题/居中/颜色
# ============================================================
ws4 = wb.active

# 标题样式
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                          fill_type="solid")
center = Alignment(horizontal="center", vertical="center")

for col in range(1, 4):
    cell = ws4.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center

# 调整列宽
ws4.column_dimensions["A"].width = 12
ws4.column_dimensions["B"].width = 12
ws4.column_dimensions["C"].width = 12

wb.save("实验数据_带样式.xlsx")
print("已保存带样式的表格")

# ============================================================
# 五、格式化数字
# ============================================================
from openpyxl.utils import get_column_letter

ws5 = wb.active
for row in range(2, ws5.max_row + 1):
    cell = ws5.cell(row=row, column=3)
    cell.number_format = "0.000"    # 3位小数显示

wb.save("实验数据_格式化.xlsx")

# ============================================================
# 六、常用操作速查
# ============================================================
# 写入：ws["A1"]=值 / ws.append([...])
# 读取：ws["A1"].value / ws.cell(r,c).value / iter_rows(values_only=True)
# 表管理：wb.sheetnames / wb.create_sheet / del wb[名]
# 样式：Font(bold=True) / PatternFill / Alignment
# 列宽：ws.column_dimensions["A"].width
# 数字格式：cell.number_format = "0.00"
# 保存：wb.save("文件名.xlsx")

# ============================================================
# 七、总结
# ============================================================
# openpyxl = 程序化操作 Excel（创建/读取/样式/格式化）
# 适合：批量生成报告表格、读取仪器导出的 xlsx
# 注意：.xls 老格式不支持，需要另存为 .xlsx

# ============================================================
# 六、易错点汇总
# ============================================================
# 1. openpyxl 只支持 .xlsx（新格式），.xls 老格式不支持
#    （要另存为 .xlsx 或用 xlrd 读 .xls）
# 2. 概念分层：Workbook(工作簿) → Worksheet(工作表) → Cell(单元格)
#    wb["Sheet1"] 取工作表，ws["A1"] 取单元格
# 3. 读文件用 load_workbook；新建用 Workbook()
# 4. 修改单元格后必须 wb.save(路径) 才落盘，
#    忘记保存 = 白改
# 5. 公式/样式是 openpyxl 的强项，但复杂样式要先了解 API
# 6. 遍历用 ws.iter_rows()，别用 A1/Z999 硬编码

# ============================================================
# 七、自测（40%基础 + 40%中等 + 20%挑战）
# ============================================================
# 【基础】
# 1. openpyxl 支持什么格式？.xls 呢？
# 2. Workbook / Worksheet / Cell 的关系？
# 3. 改完文件要调用什么才保存？
#
# 【中等】
# 4. 用 openpyxl 新建一个表，写入两行数据并保存。
# 5. 读取现有 xlsx 的某个工作表内容。
# 6. 给单元格设置字体加粗/背景色。
#
# 【挑战】
# 7. 用 iter_rows 遍历工作表并计算一列的和。
# 8. 写一个函数：把实验数据写入带样式的 Excel 报告。
