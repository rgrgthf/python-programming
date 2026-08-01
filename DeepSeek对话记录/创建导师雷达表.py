# -*- coding: utf-8 -*-
"""生成「导师雷达表-AI药学方向.xlsx」并保存到 WPS 云盘同步目录。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\22239\WPSDrive\1242849823\WPS云盘\导师雷达表-AI药学方向.xlsx"

wb = Workbook()

# ---------------- Sheet 1：导师雷达表 ----------------
ws = wb.active
ws.title = "导师雷达表"

headers = ["平台", "导师", "研究方向", "近三年代表作（待补充）", "招生意向/名额", "与我匹配度", "备注 / 下一步行动"]
rows = [
    ["中山大学（药学院）", "史滔达", "人工智能辅助的药物化学研究",
     "待查（读近2-3年论文）", "2026年：学硕 1 个名额（以当年为准）",
     "★★★★★ 最契合", "首选目标导师：持续关注招生变化；提前精读其论文，准备高质量提问"],
    ["中山大学", "王俊卿", "深度学习等进行生物大分子结构与功能性设计",
     "待查", "待查",
     "★★★☆", "备选：偏计算/结构生物学，门槛较高，先作了解"],
    ["南方医科大学（本校）", "赵培亮", "计算机辅助药物设计（CADD）",
     "待查", "待查",
     "★★★★", "预备队/练兵场：进组门槛低，可尽早接触真实科研、用 Python 做数据分析"],
    ["", "", "", "", "", "", ""],
    ["", "", "", "", "", "", ""],
]

header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

for r, row in enumerate(rows, 2):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border

widths = [20, 12, 34, 26, 26, 16, 46]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 22

# ---------------- Sheet 2：行动清单 ----------------
ws2 = wb.create_sheet("行动清单")
acts = [
    "每学期更新一次本表（导师池保持 3-5 人，随时调整）",
    "精读史滔达近 2-3 年论文：先读摘要和引言，再精读 1-2 篇",
    "关注中大药学院 / 中山药创院 / 上海药物所「优秀大学生夏令营」（大三暑假是关键窗口）",
    "大二大三主动联系本校赵培亮老师，争取进组打杂 / 用 Python 做实验数据分析",
    "持续积累 Python + AI 作品集（即战力证明，复试最硬弹药）",
    "保持绩点前 30%（大二考回卓药班 + 为保研/复试留余地）",
    "留意新出现的 AI+药学 方向导师（领域更新快，导师池要活水）",
]
ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 100
ws2.cell(row=1, column=1, value="√").font = Font(bold=True)
ws2.cell(row=1, column=2, value="行动").font = Font(bold=True, size=12)
for i, a in enumerate(acts, 2):
    ws2.cell(row=i, column=1, value="")
    ws2.cell(row=i, column=2, value="[ ] " + a).alignment = Alignment(vertical="center", wrap_text=True)

wb.save(OUT)
print("已保存:", OUT)
