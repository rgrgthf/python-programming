# -*- coding: utf-8 -*-
"""读取 WPS 学校端账号里的学习记录表与读书笔记。"""
import re
import zipfile
from openpyxl import load_workbook

BASE = r"C:\Users\22239\WPSDrive\1845681066\WPS企业云盘\南方医科大学\我的企业文档"

def read_docx(path):
    """从 .docx 中提取纯文本（docx 本质是 zip，读 word/document.xml）。"""
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8", errors="ignore")
    # 段落分隔
    xml = xml.replace("</w:p>", "\n")
    # 提取所有文本节点
    texts = re.findall(r"<w:t[^>]*>(.*?)</w:t>", xml, re.S)
    return "".join(texts)

print("=" * 60)
print("【学习记录表.xlsx】")
print("=" * 60)
wb = load_workbook(fr"{BASE}\学习记录表.xlsx", data_only=True)
print("工作表:", wb.sheetnames)
ws = wb.active
print("尺寸:", ws.dimensions)
for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
    # 只显示非空行，截断超长单元格
    vals = [("" if v is None else str(v))[:18] for v in row]
    if any(v.strip() for v in vals):
        print(" | ".join(vals))

print()
print("=" * 60)
print("【读书心得】")
print("=" * 60)
for name in ["1-幸运兔脚.docx", "2-罗夏测验.docx"]:
    path = fr"{BASE}\读书心得\献给阿尔吉侬的花束\{name}"
    try:
        text = read_docx(path)
        print(f"\n----- {name} -----")
        print(text[:1200])
    except Exception as e:
        print(f"{name} 读取失败: {e}")
